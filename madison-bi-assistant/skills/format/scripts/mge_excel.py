"""
MGE Excel Workbook Builder — Branded, print-ready Excel workbooks for Madison Group Enterprises.

This module provides helper functions that create fully branded Excel workbooks
using openpyxl. It imports all visual constants from mge_brand.py and handles
styling boilerplate so the calling code focuses only on content.

Architecture:
    mge_brand.py  — pure data: colours, fonts, metrics, formats (no openpyxl)
    mge_excel.py  — this file: openpyxl styles, builders, page setup

Entry point:
    wb, ws = create_branded_workbook("Report Title", subtitle="March 2026")
    # wb has all NamedStyles registered; ws is the active sheet, configured.

Design references:
    - Enterprise design specification Part F (library architecture)
    - openpyxl styling reference (NamedStyle, number format colour gotcha)
    - openpyxl charts & print reference (fitToPage gotcha, chart.style=None)

TASK-029, SPEC-007 — MGE Format Skill Overhaul.
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Any, Optional, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

# Brand constants — single source of truth lives in mge_brand.py.
# We import the module to keep the namespace explicit (brand.CONNECT_GREY, etc.).
# Supports both package-relative import and direct same-directory import.
try:
    from . import mge_brand as brand
except ImportError:
    import mge_brand as brand  # type: ignore[no-redef]


# ─── Internal Helpers ────────────────────────────────────────────────────────


def _font(level: brand.TypeLevel, colour_override: Optional[brand.Colour] = None) -> Font:
    """Build an openpyxl Font from a brand TypeLevel."""
    c = colour_override or level.colour
    return Font(name=brand.FONT_NAME, size=level.size_pt, bold=level.bold, color=c.hex)


def _fill_solid(colour: brand.Colour) -> PatternFill:
    """Build a solid PatternFill from a brand Colour."""
    return PatternFill("solid", fgColor=colour.hex)


def _side(colour: brand.Colour, style: str = "thin") -> Side:
    """Build a border Side."""
    return Side(style=style, color=colour.hex)


def _no_fill() -> PatternFill:
    return PatternFill(fill_type=None)


def _no_border() -> Border:
    return Border()


# ─── NamedStyle Registry ────────────────────────────────────────────────────
#
# Every NamedStyle is built once per workbook at creation time. Cells then
# reference them by name string — fast writes, small file size.
#
# NamedStyles are IMMUTABLE after registration. All properties must be
# finalised before add_named_style(). Duplicate names raise ValueError,
# so _register_styles checks before adding.


def _make_styles() -> list[NamedStyle]:
    """Build the full set of MGE NamedStyles. Returns list for registration."""

    # Shared components
    fill_none = _no_fill()
    fill_cg = _fill_solid(brand.CONNECT_GREY)
    fill_cg10 = _fill_solid(brand.CONNECT_GREY_10)
    fill_cg20 = _fill_solid(brand.CONNECT_GREY_20)
    fill_satin = _fill_solid(brand.SATIN_GREY)
    border_none = _no_border()

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # Subtotal border: single thin top line in CG
    border_subtotal = Border(top=_side(brand.CONNECT_GREY, "thin"))
    # Grand total border: double bottom line in CG
    border_grand = Border(bottom=_side(brand.CONNECT_GREY, "double"))

    styles: list[NamedStyle] = []

    def _ns(name, font, fill, alignment, number_format="General", border=None):
        s = NamedStyle(name=name)
        s.font = font
        s.fill = fill
        s.alignment = alignment
        s.number_format = number_format
        s.border = border or border_none
        styles.append(s)
        return s

    # Title / subtitle
    _ns("mge_title",
        _font(brand.EXCEL_TITLE), fill_none, align_left)
    _ns("mge_subtitle",
        _font(brand.EXCEL_SUBTITLE), fill_none, align_left)

    # Header row — white on Connect Grey
    _ns("mge_header",
        _font(brand.EXCEL_HEADER), fill_cg, align_center)

    # Subheader — CG text on CG 20% fill
    _ns("mge_subheader",
        Font(name=brand.FONT_NAME, size=9, bold=True, color=brand.CONNECT_GREY.hex),
        fill_cg20, align_left)

    # ── Body styles (white rows) ─────────────────────────────────────────
    body_font = _font(brand.EXCEL_BODY)

    _ns("mge_body", body_font, fill_none, align_left)
    _ns("mge_body_num", body_font, fill_none, align_right, brand.NUM_FMT_INTEGER)
    _ns("mge_body_currency", body_font, fill_none, align_right, brand.NUM_FMT_CURRENCY)
    _ns("mge_body_pct", body_font, fill_none, align_right, brand.NUM_FMT_PCT)
    _ns("mge_body_variance_pct", body_font, fill_none, align_right, brand.NUM_FMT_VARIANCE_PCT)
    _ns("mge_body_date", body_font, fill_none, align_left, brand.NUM_FMT_DATE)

    # ── Alternating styles (CG 10% zebra rows) ──────────────────────────
    _ns("mge_alt", body_font, fill_cg10, align_left)
    _ns("mge_alt_num", body_font, fill_cg10, align_right, brand.NUM_FMT_INTEGER)
    _ns("mge_alt_currency", body_font, fill_cg10, align_right, brand.NUM_FMT_CURRENCY)
    _ns("mge_alt_pct", body_font, fill_cg10, align_right, brand.NUM_FMT_PCT)
    _ns("mge_alt_variance_pct", body_font, fill_cg10, align_right, brand.NUM_FMT_VARIANCE_PCT)

    # ── Total styles ─────────────────────────────────────────────────────
    total_font = _font(brand.EXCEL_TOTAL)

    _ns("mge_total", total_font, fill_satin, align_left, "General", border_subtotal)
    _ns("mge_total_num", total_font, fill_satin, align_right,
        brand.NUM_FMT_INTEGER, border_subtotal)
    _ns("mge_total_currency", total_font, fill_satin, align_right,
        brand.NUM_FMT_CURRENCY, border_subtotal)

    # Grand total — double bottom border, optional CG 10% fill
    _ns("mge_grand_total", total_font, fill_cg10, align_left, "General", border_grand)

    # Footer / notes
    _ns("mge_footer",
        _font(brand.EXCEL_FOOTER), fill_none, align_left)

    return styles


def _register_styles(wb: Workbook) -> None:
    """Register all MGE NamedStyles with the workbook. Skips duplicates."""
    existing = set(wb.named_styles)
    for style in _make_styles():
        if style.name not in existing:
            wb.add_named_style(style)
            existing.add(style.name)


# ─── Entry Point ─────────────────────────────────────────────────────────────


def create_branded_workbook(
    title: str,
    subtitle: Optional[str] = None,
    author: str = "Madison Group Enterprises",
) -> tuple[Workbook, Worksheet]:
    """
    Create a new workbook with all MGE NamedStyles registered and default sheet configured.

    Returns:
        (Workbook, active_sheet) — ready for content.
    """
    wb = Workbook()
    wb.properties.creator = author

    _register_styles(wb)

    ws = wb.active
    ws.title = "Sheet1"

    # Column A is a blank left margin
    ws.column_dimensions["A"].width = 3

    # Default row heights for the title block area
    ws.row_dimensions[1].height = brand.EXCEL_ROW_HEIGHT_SPACER
    ws.row_dimensions[2].height = brand.EXCEL_ROW_HEIGHT_TITLE
    ws.row_dimensions[3].height = brand.EXCEL_ROW_HEIGHT_DATA
    ws.row_dimensions[4].height = brand.EXCEL_ROW_HEIGHT_SPACER

    # Print setup — sensible defaults for A4 landscape
    setup_print(ws, orientation="landscape", fit_to_width=1, repeat_header_row=None)

    return wb, ws


# ─── Title Block ─────────────────────────────────────────────────────────────


def write_title_block(
    ws: Worksheet,
    title: str,
    subtitle: Optional[str] = None,
    start_row: int = 1,
    start_col: int = 2,
) -> None:
    """
    Write the standard title block: blank row, title, subtitle, blank row.

    Row layout (relative to start_row):
        start_row    : blank spacer
        start_row + 1: title (merged across ~8 columns)
        start_row + 2: subtitle (merged)
        start_row + 3: blank spacer

    Sets row heights for the title block.
    """
    title_row = start_row + 1
    subtitle_row = start_row + 2
    spacer_row = start_row + 3

    # Row heights
    ws.row_dimensions[start_row].height = brand.EXCEL_ROW_HEIGHT_SPACER
    ws.row_dimensions[title_row].height = brand.EXCEL_ROW_HEIGHT_TITLE
    ws.row_dimensions[subtitle_row].height = brand.EXCEL_ROW_HEIGHT_DATA
    ws.row_dimensions[spacer_row].height = brand.EXCEL_ROW_HEIGHT_SPACER

    # Title cell — merge across a generous range (up to col K)
    end_col = max(start_col + 7, start_col)
    ws.merge_cells(
        start_row=title_row, start_column=start_col,
        end_row=title_row, end_column=end_col,
    )
    cell = ws.cell(row=title_row, column=start_col, value=title)
    cell.style = "mge_title"

    # Subtitle
    if subtitle:
        ws.merge_cells(
            start_row=subtitle_row, start_column=start_col,
            end_row=subtitle_row, end_column=end_col,
        )
        cell = ws.cell(row=subtitle_row, column=start_col, value=subtitle)
        cell.style = "mge_subtitle"


# ─── Data Table ──────────────────────────────────────────────────────────────


# Column type keywords used for auto-detection in write_data_table.
_CURRENCY_KEYWORDS = {"value", "cost", "revenue", "cogs", "margin", "price", "amount",
                      "spend", "budget", "actual", "variance", "excess", "release",
                      "purchase", "total", "subtotal"}
_PCT_KEYWORDS = {"pct", "percent", "percentage", "%", "margin%", "rate", "yield", "share"}
_DATE_KEYWORDS = {"date", "period", "month", "day", "created", "updated", "due", "start",
                  "end", "last_sale", "lastsale"}


def _infer_column_type(header: str, sample_values: list[Any]) -> str:
    """
    Infer column type from header text and sample values.

    Returns one of: 'text', 'num', 'currency', 'pct', 'date'.
    """
    h_lower = header.lower().replace(" ", "_").replace("-", "_")

    # Explicit header keyword matching
    if any(kw in h_lower for kw in _PCT_KEYWORDS):
        return "pct"
    if any(kw in h_lower for kw in _CURRENCY_KEYWORDS):
        return "currency"
    if any(kw in h_lower for kw in _DATE_KEYWORDS):
        return "date"

    # Sample value type detection
    for v in sample_values:
        if v is None:
            continue
        if isinstance(v, (date, datetime)):
            return "date"
        if isinstance(v, float):
            # Floats between -1 and 1 with no large magnitude are likely percentages
            # but we can't be sure, so default to num
            return "num"
        if isinstance(v, int):
            return "num"
        if isinstance(v, str):
            return "text"

    return "text"


def _style_name_for(col_type: str, is_alt: bool) -> str:
    """Map a column type and zebra state to a NamedStyle name."""
    if is_alt:
        return {
            "text": "mge_alt",
            "num": "mge_alt_num",
            "currency": "mge_alt_currency",
            "pct": "mge_alt_pct",
            "variance_pct": "mge_alt_variance_pct",
            "date": "mge_alt",  # date cells use text alignment; format set explicitly
        }[col_type]
    else:
        return {
            "text": "mge_body",
            "num": "mge_body_num",
            "currency": "mge_body_currency",
            "pct": "mge_body_pct",
            "variance_pct": "mge_body_variance_pct",
            "date": "mge_body_date",
        }[col_type]


def write_data_table(
    ws: Worksheet,
    headers: Sequence[str],
    data: Sequence[Sequence[Any]],
    start_row: int,
    start_col: int = 2,
    has_totals: bool = False,
    col_types: Optional[Sequence[str]] = None,
) -> tuple[int, int]:
    """
    Write a branded data table: header row, data rows with alternating stripes.

    Args:
        ws: Target worksheet.
        headers: Column header strings.
        data: Rows of data (list of lists/tuples). Each inner sequence matches headers.
        start_row: Row number for the header row.
        start_col: Column number for the first data column (default 2 = Column B).
        has_totals: If True, the last row in data is treated as a total row.
        col_types: Optional explicit column types ('text', 'num', 'currency', 'pct', 'date').
                   If None, types are auto-detected from headers and sample values.

    Returns:
        (end_row, end_col) — the last row and column written (inclusive).
    """
    if not headers:
        return (start_row, start_col)

    num_cols = len(headers)
    end_col = start_col + num_cols - 1

    # Auto-detect column types if not provided
    if col_types is None:
        # Sample first 10 rows for type inference
        samples = [[] for _ in range(num_cols)]
        for row_data in data[:10]:
            for ci in range(min(len(row_data), num_cols)):
                samples[ci].append(row_data[ci])
        col_types = [_infer_column_type(headers[ci], samples[ci]) for ci in range(num_cols)]
    else:
        col_types = list(col_types)

    # Pad col_types if shorter than headers
    while len(col_types) < num_cols:
        col_types.append("text")

    # ── Header row ───────────────────────────────────────────────────────
    ws.row_dimensions[start_row].height = brand.EXCEL_ROW_HEIGHT_HEADER
    for ci, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + ci, value=header)
        cell.style = "mge_header"

    # ── Data rows ────────────────────────────────────────────────────────
    data_row_count = len(data)
    total_row_idx = data_row_count - 1 if has_totals and data_row_count > 0 else -1

    current_row = start_row + 1
    for ri, row_data in enumerate(data):
        is_total = (ri == total_row_idx)
        is_alt = (ri % 2 == 1) and not is_total
        ws.row_dimensions[current_row].height = brand.EXCEL_ROW_HEIGHT_DATA

        for ci in range(num_cols):
            val = row_data[ci] if ci < len(row_data) else None
            cell = ws.cell(row=current_row, column=start_col + ci, value=val)

            if is_total:
                # Total row styling
                ct = col_types[ci]
                if ct in ("currency",):
                    cell.style = "mge_total_currency"
                elif ct in ("num",):
                    cell.style = "mge_total_num"
                else:
                    cell.style = "mge_total"
            else:
                # Regular data row
                style_name = _style_name_for(col_types[ci], is_alt)
                cell.style = style_name

                # Date columns in alt rows need the date format applied explicitly
                # because mge_alt doesn't carry a date number format.
                if col_types[ci] == "date":
                    cell.number_format = brand.NUM_FMT_DATE
                    # Date columns also in alt rows need CG 10% fill
                    if is_alt:
                        cell.fill = _fill_solid(brand.CONNECT_GREY_10)

        current_row += 1

    end_row = current_row - 1
    return (end_row, end_col)


# ─── Total Rows ──────────────────────────────────────────────────────────────


def write_total_row(
    ws: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    label: str,
    style: str = "subtotal",
) -> None:
    """
    Write a total row with accounting-convention borders.

    Args:
        ws: Target worksheet.
        row: Row number for the total row.
        start_col: First column.
        end_col: Last column (inclusive).
        label: Text for the label column (start_col).
        style: 'subtotal' (single top border, satin fill) or 'grand' (double bottom border).
    """
    ws.row_dimensions[row].height = brand.EXCEL_ROW_HEIGHT_DATA

    if style == "grand":
        border = Border(bottom=_side(brand.CONNECT_GREY, "double"))
        fill = _fill_solid(brand.CONNECT_GREY_10)
    else:
        border = Border(top=_side(brand.CONNECT_GREY, "thin"))
        fill = _fill_solid(brand.SATIN_GREY)

    total_font = _font(brand.EXCEL_TOTAL)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = total_font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Label cell is left-aligned
    label_cell = ws.cell(row=row, column=start_col, value=label)
    label_cell.font = total_font
    label_cell.fill = fill
    label_cell.border = border
    label_cell.alignment = Alignment(horizontal="left", vertical="center")


# ─── KPI Cards ───────────────────────────────────────────────────────────────


def write_kpi_cards(
    ws: Worksheet,
    metrics: Sequence[dict],
    start_row: int,
    start_col: int = 2,
) -> int:
    """
    Write KPI card blocks across the worksheet.

    Each metric dict: {
        'label': str,           # e.g. "Revenue"
        'value': Any,           # e.g. 1234567 or "$1.2M"
        'change': float|str,    # e.g. 0.124 or "+12.4%"
        'change_label': str,    # e.g. "vs PY"
    }

    Cards occupy 3 columns each with 1 column gap. Layout:
        Row 0: value (big number, merged 3 cols)
        Row 1: label (merged 3 cols)
        Row 2: trend indicator + change (merged 3 cols)

    Returns the row after the last KPI card row.
    """
    if not metrics:
        return start_row

    card_width = 3   # columns per card
    gap_width = 1    # columns between cards
    card_height = 4  # rows per card (value, label, change, spacer)

    big_font = Font(
        name=brand.FONT_NAME, size=20, bold=True, color=brand.CONNECT_GREY.hex,
    )
    label_font = Font(
        name=brand.FONT_NAME, size=9, color=brand.CONNECT_GREY_50.hex,
    )
    change_font_pos = Font(
        name=brand.FONT_NAME, size=10, bold=True, color=brand.CONNECT_GREY.hex,
    )
    change_font_neg = Font(
        name=brand.FONT_NAME, size=10, bold=True, color=brand.ACCENT_RED.hex,
    )
    change_font_flat = Font(
        name=brand.FONT_NAME, size=10, color=brand.CONNECT_GREY_50.hex,
    )

    align_center = Alignment(horizontal="center", vertical="center")

    for mi, m in enumerate(metrics):
        col = start_col + mi * (card_width + gap_width)
        end_merge_col = col + card_width - 1

        # Row heights
        ws.row_dimensions[start_row].height = 36
        ws.row_dimensions[start_row + 1].height = 16
        ws.row_dimensions[start_row + 2].height = 20
        ws.row_dimensions[start_row + 3].height = brand.EXCEL_ROW_HEIGHT_SPACER

        # Value row (big number)
        ws.merge_cells(
            start_row=start_row, start_column=col,
            end_row=start_row, end_column=end_merge_col,
        )
        val_cell = ws.cell(row=start_row, column=col, value=m.get("value"))
        val_cell.font = big_font
        val_cell.alignment = align_center

        # Apply currency format if value is numeric
        if isinstance(m.get("value"), (int, float)):
            val_cell.number_format = brand.NUM_FMT_CURRENCY_DYNAMIC

        # Label row
        ws.merge_cells(
            start_row=start_row + 1, start_column=col,
            end_row=start_row + 1, end_column=end_merge_col,
        )
        lbl_cell = ws.cell(row=start_row + 1, column=col, value=m.get("label", ""))
        lbl_cell.font = label_font
        lbl_cell.alignment = align_center

        # Change / trend row
        ws.merge_cells(
            start_row=start_row + 2, start_column=col,
            end_row=start_row + 2, end_column=end_merge_col,
        )
        change_val = m.get("change")
        change_label = m.get("change_label", "")

        # Determine trend direction and format
        if isinstance(change_val, (int, float)):
            if change_val > 0:
                trend_text = f"{brand.TREND_UP} +{change_val:.1%} {change_label}".strip()
                trend_font = change_font_pos
            elif change_val < 0:
                trend_text = f"{brand.TREND_DOWN} {change_val:.1%} {change_label}".strip()
                trend_font = change_font_neg
            else:
                trend_text = f"{brand.TREND_FLAT} 0.0% {change_label}".strip()
                trend_font = change_font_flat
        elif isinstance(change_val, str):
            # Pre-formatted string — guess direction from content
            if change_val.startswith("-") or change_val.startswith(brand.TREND_DOWN):
                trend_font = change_font_neg
            elif change_val.startswith("+") or change_val.startswith(brand.TREND_UP):
                trend_font = change_font_pos
            else:
                trend_font = change_font_flat
            trend_text = f"{change_val} {change_label}".strip()
        else:
            trend_text = change_label
            trend_font = change_font_flat

        chg_cell = ws.cell(row=start_row + 2, column=col, value=trend_text)
        chg_cell.font = trend_font
        chg_cell.alignment = align_center

    return start_row + card_height


# ─── Column Width ────────────────────────────────────────────────────────────


def auto_fit_columns(
    ws: Worksheet,
    start_col: int,
    end_col: int,
    start_row: int,
    end_row: int,
    min_width: float = 8.0,
    max_width: float = 50.0,
) -> None:
    """
    Set column widths based on cell content using Arial font metrics.

    Uses the per-character width table from mge_brand.py. Accounts for bold
    weight, number format expansion (e.g. "$" and "," in currency), and adds
    padding for filter dropdown arrows.

    Args:
        ws: Target worksheet.
        start_col: First column to auto-fit.
        end_col: Last column (inclusive).
        start_row: First row to scan for content.
        end_row: Last row to scan.
        min_width: Minimum column width in character units.
        max_width: Maximum column width in character units.
    """
    # Filter dropdown adds ~2.5 char widths for the arrow
    filter_padding = 2.5

    for col_idx in range(start_col, end_col + 1):
        max_w = min_width
        col_letter = get_column_letter(col_idx)

        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue

            # Get the display string. For formatted numbers, approximate the
            # formatted output width by applying the number format pattern.
            display = _format_for_width(cell.value, cell.number_format)

            is_bold = cell.font and cell.font.bold
            font_size = (cell.font.size if cell.font and cell.font.size else 9.0)

            w = brand.estimate_text_width(
                display,
                font_size=font_size,
                bold=bool(is_bold),
                padding_chars=2.0,
            )

            # Add filter dropdown allowance for header rows
            if row_idx == start_row:
                w += filter_padding

            if w > max_w:
                max_w = w

        width = max(min_width, min(max_w, max_width))
        ws.column_dimensions[col_letter].width = width


def _format_for_width(value: Any, number_format: str) -> str:
    """
    Approximate the display string for width calculation.

    Converts values to their formatted representation so width estimation
    accounts for currency symbols, commas, percentage signs, etc.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (date, datetime)):
        return value.strftime("%d/%m/%Y")

    # For numbers, simulate common format expansions
    nf = number_format or "General"

    if isinstance(value, (int, float)):
        if nf == "General":
            return str(value)

        # Currency formats
        if "$" in nf:
            if "," in nf:
                formatted = f"${abs(value):,.0f}"
            else:
                formatted = f"${abs(value):.0f}"
            if value < 0:
                formatted = f"({formatted})"
            return formatted

        # Percentage
        if "%" in nf:
            return f"{abs(value) * 100:.1f}%"

        # Integer with commas
        if nf == "#,##0" or "," in nf:
            return f"{abs(value):,.0f}"

        # Date format
        if "DD" in nf or "MM" in nf or "YY" in nf:
            return "DD/MM/YYYY"  # placeholder width

        return str(value)

    return str(value)


# ─── Negative Highlighting ───────────────────────────────────────────────────


def apply_negative_highlighting(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    cols: Sequence[int],
) -> None:
    """
    Scan specified columns and apply Accent Red font to cells with negative values.

    This applies the brand red as a cell-level font colour override — never via
    [Red] in number format strings, which would use Excel's built-in red instead
    of brand Accent Red #CF152D.

    Args:
        ws: Target worksheet.
        start_row: First data row to scan.
        end_row: Last data row (inclusive).
        cols: Column indices to check.
    """
    red_colour = brand.ACCENT_RED.hex

    for row_idx in range(start_row, end_row + 1):
        for col_idx in cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0:
                # Copy existing font and override colour to Accent Red.
                # Font is immutable — must construct a new one.
                f = cell.font
                cell.font = Font(
                    name=f.name or brand.FONT_NAME,
                    size=f.size or 9,
                    bold=f.bold,
                    italic=f.italic,
                    underline=f.underline,
                    color=red_colour,
                )


# ─── Print Setup ─────────────────────────────────────────────────────────────


def setup_print(
    ws: Worksheet,
    orientation: str = "landscape",
    fit_to_width: int = 1,
    fit_to_height: int = 0,
    repeat_header_row: Optional[int] = None,
    paper_size: int = 9,
) -> None:
    """
    Configure A4 print setup with margins, orientation, and fit-to-page.

    CRITICAL GOTCHA: fitToPage requires explicit activation via
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True).
    Setting fitToWidth alone does nothing without this.

    Args:
        ws: Target worksheet.
        orientation: 'landscape' or 'portrait'.
        fit_to_width: Number of pages wide (0 = unconstrained).
        fit_to_height: Number of pages tall (0 = unconstrained).
        repeat_header_row: Row number to repeat on every printed page (e.g. 5).
        paper_size: OOXML paper size code (9 = A4).
    """
    # Margins (inches)
    ws.page_margins = PageMargins(
        left=brand.EXCEL_MARGIN_LEFT,
        right=brand.EXCEL_MARGIN_RIGHT,
        top=brand.EXCEL_MARGIN_TOP,
        bottom=brand.EXCEL_MARGIN_BOTTOM,
        header=brand.EXCEL_MARGIN_HEADER,
        footer=brand.EXCEL_MARGIN_FOOTER,
    )

    # Page setup
    ws.page_setup.paperSize = paper_size
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = fit_to_height

    # THE GOTCHA: activate fitToPage. Without this line, fit-to-width is ignored.
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # Centre content on page horizontally
    ws.print_options.horizontalCentered = True

    # Repeat header row on every printed page
    if repeat_header_row is not None:
        ws.print_title_rows = f"{repeat_header_row}:{repeat_header_row}"


# ─── Freeze Panes ───────────────────────────────────────────────────────────


def freeze_panes(ws: Worksheet, row: int, col: int = 2) -> None:
    """
    Freeze panes at the specified row and column.

    Default freezes below the given row and to the right of Column B
    (after blank Column A margin). The freeze cell is the first non-frozen cell.

    Args:
        ws: Target worksheet.
        row: Freeze below this row (row itself is frozen/visible).
        col: Freeze to the right of this column.
    """
    cell_ref = f"{get_column_letter(col)}{row + 1}"
    ws.freeze_panes = cell_ref


# ─── Branded Charts ──────────────────────────────────────────────────────────

# Chart type mapping for convenience
_CHART_TYPES = {
    "bar": BarChart,
    "column": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
}


def add_branded_chart(
    ws: Worksheet,
    chart_type: str,
    data_ref: Reference,
    cat_ref: Optional[Reference],
    anchor: str,
    title: Optional[str] = None,
    width: float = 18,
    height: float = 10,
    titles_from_data: bool = True,
    legend_position: str = "b",
    y_axis_title: Optional[str] = None,
    x_axis_title: Optional[str] = None,
    grouping: Optional[str] = None,
    series_colours: Optional[Sequence[brand.Colour]] = None,
    **kwargs,
) -> None:
    """
    Create a chart with brand colours, no default Excel theme leakage.

    Args:
        ws: Target worksheet.
        chart_type: One of 'bar', 'column', 'line', 'pie', 'area'.
        data_ref: openpyxl Reference for data series.
        cat_ref: openpyxl Reference for category labels (x-axis). None for pie.
        anchor: Cell reference for chart placement, e.g. "B20".
        title: Chart title string. None for no title.
        width: Chart width in cm.
        height: Chart height in cm.
        titles_from_data: If True, first row of data_ref contains series names.
        legend_position: 'b' (bottom), 't', 'r', 'l', or None to hide.
        y_axis_title: Y-axis label. None for no label.
        x_axis_title: X-axis label. None for no label.
        grouping: Chart grouping — 'clustered', 'stacked', 'percentStacked'.
        series_colours: Brand colours for series. Defaults to CHART_COLOURS.
        **kwargs: Additional properties set on the chart object.
    """
    chart_cls = _CHART_TYPES.get(chart_type.lower())
    if chart_cls is None:
        raise ValueError(f"Unsupported chart type: {chart_type}. Use one of {list(_CHART_TYPES)}")

    chart = chart_cls()

    # CRITICAL: Use style 2 (minimal built-in) then override series colours.
    # chart.style = None can produce invalid XML that triggers Excel recovery.
    # Style 2 is the most basic style; our series colour loop below replaces all
    # theme colours, so the only purpose of the style is structural validity.
    chart.style = 2

    # Dimensions
    chart.width = width
    chart.height = height

    # Bar chart specifics
    if chart_type.lower() in ("bar", "column"):
        chart.type = "col" if chart_type.lower() == "column" else "bar"
        if grouping:
            chart.grouping = grouping
            if grouping in ("stacked", "percentStacked"):
                chart.overlap = 100

    # Area chart grouping
    if chart_type.lower() == "area" and grouping:
        chart.grouping = grouping

    # Add data
    chart.add_data(data_ref, titles_from_data=titles_from_data)
    if cat_ref is not None:
        chart.set_categories(cat_ref)

    # Title
    if title:
        chart.title = title
        # Brand the title font
        _apply_chart_title_font(chart, title)

    # Legend
    if legend_position is None:
        chart.legend = None
    elif chart.legend is not None:
        chart.legend.position = legend_position

    # Axis titles
    if y_axis_title and hasattr(chart, "y_axis"):
        chart.y_axis.title = y_axis_title
    if x_axis_title and hasattr(chart, "x_axis"):
        chart.x_axis.title = x_axis_title

    # Subtle gridlines — CG 10% at 0.5pt
    if hasattr(chart, "y_axis") and chart.y_axis.majorGridlines is not None:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill=brand.CONNECT_GREY_10.hex, w=6350)
        chart.y_axis.majorGridlines.graphicalProperties = gp

    # Make plot area transparent
    try:
        if chart.plot_area is not None and chart.plot_area.graphicalProperties is not None:
            chart.plot_area.graphicalProperties.noFill = True
    except AttributeError:
        pass  # Some chart types don't initialise plot_area graphicalProperties

    # Remove chart border
    try:
        chart.graphical_properties.line.noFill = True
    except AttributeError:
        pass  # Safe fallback if graphical_properties isn't fully initialised

    # Apply kwargs
    for k, v in kwargs.items():
        setattr(chart, k, v)

    # ── Colour every series with brand palette ───────────────────────────
    colours = series_colours or brand.CHART_COLOURS
    for si, series in enumerate(chart.series):
        colour = colours[si % len(colours)]
        series.graphicalProperties.solidFill = colour.hex
        # For line charts, also set line colour and width
        if chart_type.lower() == "line":
            series.graphicalProperties.line.solidFill = colour.hex
            series.graphicalProperties.line.width = 25400  # 2pt in EMU
            series.graphicalProperties.noFill = True  # No area fill under line
        # For pie charts, colour each slice individually
        if chart_type.lower() == "pie":
            for di in range(len(chart.series[0].val.numRef.numCache.pt) if hasattr(chart.series[0].val, 'numRef') else 0):
                pt = DataPoint(idx=di)
                c = colours[di % len(colours)]
                pt.graphicalProperties.solidFill = c.hex
                series.data_points.append(pt)

    ws.add_chart(chart, anchor)


def _apply_chart_title_font(chart, title_text: str) -> None:
    """Apply brand font to chart title (Connect Grey, 11pt bold Arial)."""
    try:
        cp = CharacterProperties(
            latin=DrawingFont(typeface=brand.FONT_NAME),
            sz=1100,  # 11pt in hundredths
            b=True,
            solidFill=brand.CONNECT_GREY.hex,
        )
        paragraph = Paragraph(
            pPr=ParagraphProperties(defRPr=cp),
            endParaRPr=cp,
        )
        chart.title.txPr = RichText(p=[paragraph])
    except (AttributeError, TypeError):
        # Fallback: title string is still set, just without custom font
        pass


# ─── AutoFilter ──────────────────────────────────────────────────────────────


def add_autofilter(ws: Worksheet, start_row: int, start_col: int, end_col: int) -> None:
    """
    Apply autofilter to the header row spanning start_col to end_col.

    Args:
        ws: Target worksheet.
        start_row: The header row number.
        start_col: First column of the filter range.
        end_col: Last column of the filter range.
    """
    start_ref = f"{get_column_letter(start_col)}{start_row}"
    # Find last row with data by scanning down from header
    end_row = start_row
    for row in ws.iter_rows(min_row=start_row + 1, min_col=start_col,
                            max_col=start_col, values_only=False):
        if row[0].value is not None:
            end_row = row[0].row
        else:
            break

    # If we didn't find any data, just use header row + 1
    if end_row == start_row:
        end_row = start_row + 1

    end_ref = f"{get_column_letter(end_col)}{end_row}"
    ws.auto_filter.ref = f"{start_ref}:{end_ref}"


# ─── Convenience: Full Table Pipeline ────────────────────────────────────────


def build_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: Optional[str],
    headers: Sequence[str],
    data: Sequence[Sequence[Any]],
    col_types: Optional[Sequence[str]] = None,
    has_totals: bool = False,
    tab_colour: Optional[str] = None,
    freeze_row: Optional[int] = None,
    negative_cols: Optional[Sequence[int]] = None,
    orientation: str = "landscape",
    repeat_header_row: bool = True,
    autofilter: bool = True,
    auto_fit: bool = True,
) -> Worksheet:
    """
    High-level convenience: create a complete branded sheet with title block, data
    table, auto-fitted columns, print setup, freeze panes, and autofilter.

    This is the most common workflow — call this instead of composing individual
    functions for standard data tables.

    Args:
        wb: Target workbook (must have styles registered via create_branded_workbook).
        sheet_name: Tab name.
        title: Title text.
        subtitle: Subtitle text.
        headers: Column headers.
        data: Data rows.
        col_types: Optional column type hints.
        has_totals: Whether the last data row is a total.
        tab_colour: Hex colour for the sheet tab. Default Connect Grey.
        freeze_row: Row to freeze below. Default auto-detects (header row).
        negative_cols: Column indices for negative highlighting.
        orientation: Print orientation.
        repeat_header_row: Whether to repeat the header on printed pages.
        autofilter: Whether to add autofilter.
        auto_fit: Whether to auto-fit column widths.

    Returns:
        The created worksheet.
    """
    ws = wb.create_sheet(title=sheet_name)

    # Column A margin
    ws.column_dimensions["A"].width = 3

    # Title block (rows 1-4)
    write_title_block(ws, title, subtitle)

    # Data table starts at row 5
    header_row = brand.EXCEL_HEADER_ROW  # 5
    end_row, end_col = write_data_table(
        ws, headers, data, start_row=header_row, col_types=col_types,
        has_totals=has_totals,
    )

    # Auto-fit columns
    if auto_fit:
        auto_fit_columns(ws, brand.EXCEL_LABEL_COL, end_col, header_row, end_row)

    # Negative highlighting
    if negative_cols:
        data_start = header_row + 1
        data_end = end_row - 1 if has_totals else end_row
        apply_negative_highlighting(ws, data_start, data_end, negative_cols)

    # Print setup
    setup_print(
        ws,
        orientation=orientation,
        fit_to_width=1,
        repeat_header_row=header_row if repeat_header_row else None,
    )

    # Freeze panes — below header row, after label column
    if freeze_row is not None:
        freeze_panes(ws, freeze_row)
    else:
        freeze_panes(ws, header_row)

    # Autofilter
    if autofilter and data:
        add_autofilter(ws, header_row, brand.EXCEL_LABEL_COL, end_col)

    # Tab colour
    ws.sheet_properties.tabColor = tab_colour or brand.TAB_COLOUR_DATA

    return ws


# ─── Module version ──────────────────────────────────────────────────────────

__version__ = "1.0.0"
